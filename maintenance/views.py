# REV: 1.0 | 2025-09-24 | Hash: ed39944f | Parça: 1/1
# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import io
from datetime import date
from typing import Dict, List, Tuple, Iterable

from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import permission_required
from django.core.exceptions import ValidationError
from django.db.models import OuterRef, Subquery, DateField
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.utils.dateparse import parse_date

from .models import CalibrationAsset, CalibrationRecord

# ---------------------------
# Yardımcılar
# ---------------------------

def _safe_str(val):
    return "" if val is None else str(val).strip()

def _parse_bool(val) -> bool:
    s = _safe_str(val).lower()
    if s in ("1", "true", "evet", "yes", "y", "on", "aktif"):
        return True
    if s in ("0", "false", "hayır", "hayir", "no", "n", "off", "pasif"):
        return False
    return True  # boş ise aktif varsayalım

def _parse_date_any(val):
    if isinstance(val, date):
        return val
    s = _safe_str(val)
    if not s:
        return None
    # YYYY-MM-DD gibi formatları yakalar
    d = parse_date(s)
    if d:
        return d
    # Gün/Ay/Yıl varyasyonlarını kaba şekilde ele al
    for sep in (".", "/", "-"):
        parts = s.split(sep)
        if len(parts) == 3:
            # dene: DD MM YYYY ya da YYYY MM DD
            a, b, c = parts
            # YYYY-first
            d = parse_date(f"{a.zfill(4)}-{b.zfill(2)}-{c.zfill(2)}")
            if d:
                return d
            # DD-first
            d = parse_date(f"{c.zfill(4)}-{b.zfill(2)}-{a.zfill(2)}")
            if d:
                return d
    return None

# Sütun->Model alan haritası (varsayılan eşleşme isimleri)
DEFAULT_FIELD_MAP = {
    "asset_code": ["cihaz kodu", "asset code", "code", "kodu"],
    "asset_name": ["cihaz adı", "adı", "name", "equipment"],
    "location": ["lokasyon", "konum", "location", "bölüm", "hat"],
    "brand": ["marka", "brand"],
    "model": ["model", "tip"],
    "serial_no": ["seri no", "serino", "serial", "serial no", "serial number"],
    "measure_range": ["ölçüm aralığı", "aralık", "range", "measure range"],
    "resolution": ["çözünürlük", "resolution"],
    "unit": ["birim", "unit"],
    "accuracy": ["doğruluk", "accuracy", "hassasiyet"],
    "uncertainty": ["belirsizlik", "uncertainty"],
    "acceptance_criteria": ["kabul kriteri", "acceptance", "acceptance criteria", "limit"],
    "calibration_method": ["yöntem", "prosedür", "metod", "method", "procedure"],
    "standard_device": ["referans cihaz", "standart cihaz", "standard device", "reference"],
    "standard_id": ["standart id", "standard id", "referans id"],
    "owner": ["sorumlu", "owner", "kullanıcı"],
    "responsible_email": ["e-posta", "email", "mail", "responsible email"],
    "is_active": ["aktif", "durum", "active", "status"],
}

IMPORTABLE_FIELDS = list(DEFAULT_FIELD_MAP.keys())

def _auto_guess_mapping(headers: List[str]) -> Dict[str, str]:
    """
    headers: dosyadaki sütun adları (lowercase olabilir)
    dönen: {model_field: header_name_lower}
    """
    mapping: Dict[str, str] = {}
    low_headers = {h.lower(): h.lower() for h in headers}  # değerleri de lower saklıyoruz
    for field, aliases in DEFAULT_FIELD_MAP.items():
        for alias in aliases:
            if alias in low_headers:
                mapping[field] = low_headers[alias]
                break
        if field not in mapping:
            if field in low_headers:
                mapping[field] = low_headers[field]
    return mapping

def _read_csv_preview(f) -> Tuple[List[str], List[List[str]]]:
    data = f.read()
    if isinstance(data, bytes):
        data = data.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(data))
    rows = list(reader)
    headers = rows[0] if rows else []
    preview = rows[1:21]  # ilk 20 satır
    return headers, preview

def _read_xlsx_preview(f) -> Tuple[List[str], List[List[str]]]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        raise ValidationError("'.xlsx' okumak için 'openpyxl' kurulu olmalı: pip install openpyxl")
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [(_safe_str(c)) for c in (rows[0] or [])] if rows else []
    preview = [[_safe_str(c) for c in r] for r in rows[1:21]]
    return headers, preview

def _iter_csv_rows(f) -> Iterable[Dict[str, str]]:
    data = f.read()
    if isinstance(data, bytes):
        data = data.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(data))
    for row in reader:
        yield {k: _safe_str(v) for k, v in row.items()}

def _iter_xlsx_rows(f) -> Iterable[Dict[str, str]]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        raise ValidationError("'.xlsx' okumak için 'openpyxl' kurulu olmalı: pip install openpyxl")
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    # başlık satırı
    headers = [(_safe_str(c)) for c in next(ws.iter_rows(values_only=True))]
    for row in ws.iter_rows(values_only=True):
        vals = [(_safe_str(c)) for c in row]
        yield dict(zip(headers, vals))

def _apply_mapping(row: Dict[str, str], mapping: Dict[str, str]) -> Dict[str, str]:
    """Dosya satırını model alanlarına projekte et."""
    m = {}
    for field in IMPORTABLE_FIELDS:
        header = mapping.get(field, "")
        m[field] = row.get(header, "")
    return m

def _upsert_asset(mapped: Dict[str, str]) -> Tuple[str, CalibrationAsset, bool]:
    """asset_code zorunlu; varsa update, yoksa create. returns: (status, obj, created)"""
    code = _safe_str(mapped.get("asset_code"))
    name = _safe_str(mapped.get("asset_name"))
    if not code:
        return ("SKIPPED_NO_CODE", None, False)  # type: ignore
    defaults = dict(
        asset_name=name or code,
        location=_safe_str(mapped.get("location")),
        brand=_safe_str(mapped.get("brand")),
        model=_safe_str(mapped.get("model")),
        serial_no=_safe_str(mapped.get("serial_no")),
        measure_range=_safe_str(mapped.get("measure_range")),
        resolution=_safe_str(mapped.get("resolution")),
        unit=_safe_str(mapped.get("unit")),
        accuracy=_safe_str(mapped.get("accuracy")),
        uncertainty=_safe_str(mapped.get("uncertainty")),
        acceptance_criteria=_safe_str(mapped.get("acceptance_criteria")),
        calibration_method=_safe_str(mapped.get("calibration_method")),
        standard_device=_safe_str(mapped.get("standard_device")),
        standard_id=_safe_str(mapped.get("standard_id")),
        owner=_safe_str(mapped.get("owner")),
        responsible_email=_safe_str(mapped.get("responsible_email")),
        is_active=_parse_bool(mapped.get("is_active")),
    )
    obj, created = CalibrationAsset.objects.update_or_create(asset_code=code, defaults=defaults)
    return ("CREATED" if created else "UPDATED", obj, created)

# ---------------------------
# Tablo + (isteğe bağlı) Sihirbaz
# ---------------------------

@staff_member_required
def calibration_full_table(request: HttpRequest) -> HttpResponse:
    """
    /calibration/table/  -> normal tablo
    /calibration/table/?import=1 -> içe aktarma sihirbazı
    """
    if request.GET.get("import") == "1":
        return calibration_import_wizard(request)

    # Son/gelecek tarih anotasyonları (Subquery ile)
    latest_last = CalibrationRecord.objects.filter(
        asset_id=OuterRef("pk")
    ).order_by("-last_calibration").values("last_calibration")[:1]
    latest_next = CalibrationRecord.objects.filter(
        asset_id=OuterRef("pk")
    ).order_by("-next_calibration").values("next_calibration")[:1]

    assets = (
        CalibrationAsset.objects.all()
        .annotate(last_cal=Subquery(latest_last, output_field=DateField()))
        .annotate(next_cal=Subquery(latest_next, output_field=DateField()))
        .order_by("asset_code")
    )

    ctx = {"assets": assets}
    return render(request, "maintenance/calibration_table.html", ctx)

# ---------------------------
# İçe Aktarım Sihirbazı (aynı route altında)
# ---------------------------

@staff_member_required
@permission_required("maintenance.add_calibrationasset")
def calibration_import_wizard(request: HttpRequest) -> HttpResponse:
    """
    Adımlar:
      - GET: upload formu
      - POST action=preview: dosya başlıklarını ve ilk 20 satırı önizleme + otomatik eşleme
      - POST action=import : mapping ile içe aktar ve özet döndür
    """
    ctx = {
        "headers": [],
        "preview": [],
        "mapping": {},
        "importable_fields": IMPORTABLE_FIELDS,
        "guessed": {},
        "ui_map": [],  # <— şablon için: [{'field': 'asset_code', 'selected': 'asset code'}, ...]
        "result": None,
        "error": None,
    }

    if request.method == "GET":
        return render(request, "maintenance/import_wizard.html", ctx)

    action = request.POST.get("action", "preview")
    upfile = request.FILES.get("file")

    if not upfile:
        ctx["error"] = "Lütfen bir .xlsx veya .csv dosyası seçin."
        return render(request, "maintenance/import_wizard.html", ctx)

    filename = (upfile.name or "").lower()
    is_xlsx = filename.endswith(".xlsx")
    is_csv = filename.endswith(".csv")

    try:
        # Aynı request yaşam döngüsünde dosyayı iki kez okuyabileceğimiz bir buffer
        mem = io.BytesIO(upfile.read())
        mem.seek(0)

        if action == "preview":
            if is_csv:
                headers, preview = _read_csv_preview(io.BytesIO(mem.getvalue()))
            elif is_xlsx:
                headers, preview = _read_xlsx_preview(io.BytesIO(mem.getvalue()))
            else:
                raise ValidationError("Yalnızca .xlsx veya .csv desteklenir.")

            # guessed: field -> header_lower
            guessed = _auto_guess_mapping([h.lower() for h in headers])
            ui_map = [{"field": f, "selected": guessed.get(f, "")} for f in IMPORTABLE_FIELDS]

            ctx.update({
                "headers": headers,
                "preview": preview,
                "guessed": guessed,
                "ui_map": ui_map,
            })
            return render(request, "maintenance/import_wizard.html", ctx)

        elif action == "import":
            # mapping alanlarını topla
            mapping = {}
            for field in IMPORTABLE_FIELDS:
                header = request.POST.get(f"map_{field}", "")
                if header:
                    mapping[field] = header
            if "asset_code" not in mapping:
                raise ValidationError("En azından 'asset_code' için bir sütun seçmelisiniz.")

            # satırları tara ve içe aktar
            created = updated = skipped = 0
            report: List[Tuple[str, str]] = []

            def _iter():
                if is_csv:
                    return _iter_csv_rows(io.BytesIO(mem.getvalue()))
                elif is_xlsx:
                    return _iter_xlsx_rows(io.BytesIO(mem.getvalue()))
                else:
                    raise ValidationError("Yalnızca .xlsx veya .csv desteklenir.")

            for row in _iter():
                mapped = _apply_mapping(row, mapping)
                status, obj, is_created = _upsert_asset(mapped)
                if status == "SKIPPED_NO_CODE":
                    skipped += 1
                    report.append(("SKIPPED", "asset_code boş"))
                else:
                    if is_created:
                        created += 1
                    else:
                        updated += 1
                    report.append((status, obj.asset_code))

            ctx["result"] = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "report": report[:500],
            }
            return render(request, "maintenance/import_wizard.html", ctx)

        else:
            raise ValidationError("Bilinmeyen işlem.")

    except ValidationError as ve:
        ctx["error"] = str(ve)
        return render(request, "maintenance/import_wizard.html", ctx)
    except Exception as ex:
        ctx["error"] = f"İçe aktarma sırasında beklenmeyen hata: {ex}"
        return render(request, "maintenance/import_wizard.html", ctx)
